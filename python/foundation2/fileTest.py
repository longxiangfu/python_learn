"""
file 文件操作
"""
print("---获取当前目录---")
import os
print(os.getcwd()) # D:\workspace\workspace-python\python\foundation2

# print("---w: 以覆盖写的方式打开文件。文件不存在则创建文件，文件存在则覆盖文件---")
# f = open("test.txt", "w", encoding="utf-8")
# f.write("hello world")
# f.close()
#
#
# print("---a: 以追加写的方式打开文件。文件不存在则创建文件，文件存在则追加文件---")
# f = open("test.txt", "a", encoding="utf-8")
# f.write("追加写")
# f.close()

# print("---r: 以读的方式打开文件---")
# f = open("test.txt", "r", encoding="utf-8")
# print(f.read()) # hello world追加写


# print("---with ... as f: 不用主动关闭文件---")
# with open("test.txt", "a", encoding="utf-8") as f:
#     f.write("又一个追加写")


# print("---for line in f: 逐行读取文件     f是可迭代的---")
# f = open("test.txt", "r", encoding="utf-8")
# for line in f:
#     print(line)


print("---csv工具读写csv文件---")
import csv
data = [['name', 'number'], ['python', 111], ['java', 222], ['php', 333]]
with open('csvfile.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(data)

f = open('csvfile.csv', 'r')
reader = csv.reader(f)
for line in reader:
    print(line)
# ['name', 'number']
# ['python', '111']
# ['java', '222']
# ['php', '333']


print("---excel工具读写excel文件---")
from openpyxl.workbook.workbook import Workbook
# 创建工作簿
wb = Workbook()
# 获取当前活动工作表
ws = wb.active
print(ws.title) # Sheet
ws.title = 'python'
print(ws.title) # python

# 创建工作表
ws2 = wb.create_sheet('java')
print(ws2.title) # java

# 查看所有工作表
print(wb.sheetnames) # ['python', 'java']

# 写入ws工作表
ws['F1'] = 111
ws.cell(row=2, column=5, value=222)

# 写入ws2工作表
ws2.cell(row=3, column=3, value='java')

# 保存
wb.save('excel.xlsx')


